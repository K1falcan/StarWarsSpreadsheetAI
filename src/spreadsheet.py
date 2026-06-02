from __future__ import annotations
 
import logging
from typing import Any, Callable, Dict, List, NamedTuple
 
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
 
from .swapi_client import SwapiClient
 
log = logging.getLogger(__name__)
 
# ---- styling -----------------------------------------------------------------
 
# Font sizes bumped up a notch from openpyxl's default of 11 for readability.
BODY_FONT_SIZE = 13
HEADER_FONT_SIZE = 14

HEADER_FILL = PatternFill("solid", fgColor="1F2A44")  # deep Star Wars navy
HEADER_FONT = Font(bold=True, color="FFE81F", size=HEADER_FONT_SIZE)  # the iconic yellow
BODY_FONT = Font(size=BODY_FONT_SIZE)
# Rows for the "important" entities (main characters, most-used ships, most-shown
# planets) get this blue fill. Light enough that the black body text stays legible.
HIGHLIGHT_FILL = PatternFill("solid", fgColor="9DC3E6")
HIGHLIGHT_FONT = Font(size=BODY_FONT_SIZE, bold=True)
THIN = Side(style="thin", color="D0D0D0")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(vertical="top", wrap_text=True)
HEADER_ALIGN = Alignment(vertical="center", horizontal="center", wrap_text=True)
 
MAX_COL_WIDTH = 60
MIN_COL_WIDTH = 12

# "Important" rows are highlighted when an entity appears in at least this many
# films. Thresholds differ per sheet because the distributions differ (films
# total only 6) and were chosen at the natural break in each distribution.
MAIN_CHARACTER_MIN_FILMS = 4  # -> 9 saga leads (Luke, Leia, Vader, Yoda, ...)
TOP_STARSHIP_MIN_FILMS = 3    # -> the 6 most-used ships (X-wing, Falcon, ...)
TOP_PLANET_MIN_FILMS = 3      # -> the 4 most-shown worlds (Tatooine, Naboo, ...)


def _films_at_least(minimum: int) -> Callable[[Dict[str, Any]], bool]:
    """Predicate: record appears in at least ``minimum`` films."""
    return lambda record: len(record.get("films", [])) >= minimum
 
 
class Column(NamedTuple):
    """One spreadsheet column: a header and how to derive its value from a record."""
 
    header: str
    value: Callable[[Dict[str, Any], SwapiClient], Any]
 
 
def _text(field: str) -> Callable[[Dict[str, Any], SwapiClient], Any]:
    """Column accessor for a plain scalar field."""
    return lambda record, _client: record.get(field, "")
 
 
def _refs(field: str) -> Callable[[Dict[str, Any], SwapiClient], Any]:
    """Column accessor that resolves a list of resource URLs to a names string."""
    return lambda record, client: ", ".join(client.names_for(record.get(field, [])))
 
 
def _ref(field: str) -> Callable[[Dict[str, Any], SwapiClient], Any]:
    """Column accessor that resolves a single resource URL to a name."""
    return lambda record, client: client.name_for(record[field]) if record.get(field) else ""
 
 
# ---- per-resource column definitions -----------------------------------------
 
CHARACTER_COLUMNS = [
    Column("Name", _text("name")),
    Column("Height (cm)", _text("height")),
    Column("Mass (kg)", _text("mass")),
    Column("Hair Color", _text("hair_color")),
    Column("Skin Color", _text("skin_color")),
    Column("Eye Color", _text("eye_color")),
    Column("Birth Year", _text("birth_year")),
    Column("Gender", _text("gender")),
    Column("Homeworld", _ref("homeworld")),
    Column("Species", _refs("species")),
    Column("Starships", _refs("starships")),
    Column("Vehicles", _refs("vehicles")),
    Column("Films", _refs("films")),
]
 
PLANET_COLUMNS = [
    Column("Name", _text("name")),
    Column("Rotation Period", _text("rotation_period")),
    Column("Orbital Period", _text("orbital_period")),
    Column("Diameter", _text("diameter")),
    Column("Climate", _text("climate")),
    Column("Gravity", _text("gravity")),
    Column("Terrain", _text("terrain")),
    Column("Surface Water", _text("surface_water")),
    Column("Population", _text("population")),
    Column("Residents", _refs("residents")),
    Column("Films", _refs("films")),
]
 
SPECIES_COLUMNS = [
    Column("Name", _text("name")),
    Column("Classification", _text("classification")),
    Column("Designation", _text("designation")),
    Column("Average Height", _text("average_height")),
    Column("Skin Colors", _text("skin_colors")),
    Column("Hair Colors", _text("hair_colors")),
    Column("Eye Colors", _text("eye_colors")),
    Column("Average Lifespan", _text("average_lifespan")),
    Column("Homeworld", _ref("homeworld")),
    Column("Language", _text("language")),
    Column("Members", _refs("people")),
    Column("Films", _refs("films")),
]
 
STARSHIP_COLUMNS = [
    Column("Name", _text("name")),
    Column("Model", _text("model")),
    Column("Manufacturer", _text("manufacturer")),
    Column("Cost (credits)", _text("cost_in_credits")),
    Column("Length", _text("length")),
    Column("Max Atmosphering Speed", _text("max_atmosphering_speed")),
    Column("Crew", _text("crew")),
    Column("Passengers", _text("passengers")),
    Column("Cargo Capacity", _text("cargo_capacity")),
    Column("Consumables", _text("consumables")),
    Column("Hyperdrive Rating", _text("hyperdrive_rating")),
    Column("MGLT", _text("MGLT")),
    Column("Starship Class", _text("starship_class")),
    Column("Pilots", _refs("pilots")),
    Column("Films", _refs("films")),
]
 
VEHICLE_COLUMNS = [
    Column("Name", _text("name")),
    Column("Model", _text("model")),
    Column("Manufacturer", _text("manufacturer")),
    Column("Cost (credits)", _text("cost_in_credits")),
    Column("Length", _text("length")),
    Column("Max Atmosphering Speed", _text("max_atmosphering_speed")),
    Column("Crew", _text("crew")),
    Column("Passengers", _text("passengers")),
    Column("Cargo Capacity", _text("cargo_capacity")),
    Column("Consumables", _text("consumables")),
    Column("Vehicle Class", _text("vehicle_class")),
    Column("Pilots", _refs("pilots")),
    Column("Films", _refs("films")),
]
 
FILM_COLUMNS = [
    Column("Episode", _text("episode_id")),
    Column("Title", _text("title")),
    Column("Director", _text("director")),
    Column("Producer", _text("producer")),
    Column("Release Date", _text("release_date")),
    Column("Characters", _refs("characters")),
    Column("Planets", _refs("planets")),
    Column("Species", _refs("species")),
    Column("Starships", _refs("starships")),
    Column("Vehicles", _refs("vehicles")),
]
 
 
def _write_sheet(
    ws: Worksheet,
    columns: List[Column],
    records: List[Dict[str, Any]],
    client: SwapiClient,
    highlight: Callable[[Dict[str, Any]], bool] | None = None,
) -> None:
    """Write a header row plus one row per record, then style the sheet.

    ``highlight`` is an optional predicate; rows whose record satisfies it get
    the blue highlight fill (main characters / most-used ships / most-shown planets).
    """
    headers = [c.header for c in columns]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
 
    for record in records:
        ws.append([col.value(record, client) for col in columns])
 
    highlighted = [bool(highlight(r)) for r in records] if highlight else None
    _style_body(ws, len(columns), len(records), highlighted)
 
 
def _style_body(
    ws: Worksheet,
    n_cols: int,
    n_rows: int,
    highlighted: List[bool] | None = None,
) -> None:
    """Apply font/borders/wrapping, freeze the header, add a filter, size columns."""
    for offset, row in enumerate(ws.iter_rows(min_row=2, max_row=n_rows + 1, max_col=n_cols)):
        is_highlighted = bool(highlighted[offset]) if highlighted else False
        for cell in row:
            cell.alignment = WRAP_TOP
            cell.border = CELL_BORDER
            if is_highlighted:
                cell.fill = HIGHLIGHT_FILL
                cell.font = HIGHLIGHT_FONT
            else:
                cell.font = BODY_FONT
 
    ws.freeze_panes = "A2"
    if n_rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"
 
    # Width = longest cell in the column, clamped. Multi-name reference cells can
    # get huge, so we cap them and let wrap_text handle overflow.
    for idx in range(1, n_cols + 1):
        letter = get_column_letter(idx)
        longest = max(
            (len(str(ws.cell(row=r, column=idx).value or "")) for r in range(1, n_rows + 2)),
            default=MIN_COL_WIDTH,
        )
        ws.column_dimensions[letter].width = max(MIN_COL_WIDTH, min(longest + 2, MAX_COL_WIDTH))
 
 
def build_workbook(client: SwapiClient, data: Dict[str, List[Dict[str, Any]]]) -> Workbook:
    """Build the full workbook from already-fetched swapi data."""
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
 
    films = data["films"]
    # Movie order, two ways the user asked for.
    chronological = sorted(films, key=lambda f: f.get("episode_id", 0))
    by_release = sorted(films, key=lambda f: f.get("release_date", ""))
 
    def alpha(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Sort records alphabetically by their 'name' field (case-insensitive)."""
        return sorted(records, key=lambda r: r.get("name", "").lower())
 
    # (title, columns, records, highlight-predicate-or-None). The predicate marks
    # the "important" rows that get the blue highlight.
    sheets = [
        ("Characters", CHARACTER_COLUMNS, alpha(data["people"]), _films_at_least(MAIN_CHARACTER_MIN_FILMS)),
        ("Planets", PLANET_COLUMNS, alpha(data["planets"]), _films_at_least(TOP_PLANET_MIN_FILMS)),
        ("Species", SPECIES_COLUMNS, alpha(data["species"]), None),
        ("Starships", STARSHIP_COLUMNS, alpha(data["starships"]), _films_at_least(TOP_STARSHIP_MIN_FILMS)),
        ("Vehicles", VEHICLE_COLUMNS, alpha(data["vehicles"]), None),
        ("Films (Chronological)", FILM_COLUMNS, chronological, None),
        ("Films (Release Order)", FILM_COLUMNS, by_release, None),
    ]
 
    for title, columns, records, highlight in sheets:
        ws = wb.create_sheet(title=title)
        _write_sheet(ws, columns, records, client, highlight)
        log.info("Wrote sheet %r with %d rows", title, len(records))
 
    return wb