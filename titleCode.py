from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore
 
# The Phantom Menace
phantomMenace = "The Phantom Menace"
phantomMenaceNum1 = 1      # Chronological order
phantomMenaceNum2 = 4      # Release order
 
# Attack of the Clones
attackOfTheClones = "Attack of the Clones"
attackOfTheClonesNum1 = 2  # Chronological order
attackOfTheClonesNum2 = 5  # Release order
 
# Revenge of the Sith
revengeOfTheSith = "Revenge of the Sith"
revengeOfTheSithNum1 = 3   # Chronological order
revengeOfTheSithNum2 = 6  # Release order
 
# A New Hope
newHope = "A New Hope"
newHopeNum1 = 4            # Chronological order
newHopeNum2 = 1            # Release order
 
# The Empire Strikes Back
empireStrikesBack = "The Empire Strikes Back"
empireStrikesBackNum1 = 5  # Chronological order
empireStrikesBackNum2 = 2  # Release order
 
# Return of the Jedi
returnOfTheJedi = "Return of the Jedi"
returnOfTheJediNum1 = 6    # Chronological order
returnOfTheJediNum2 = 3    # Release order
 
# The Force Awakens
forceAwakens = "The Force Awakens"
forceAwakensNum1 = 7       # Chronological order
forceAwakensNum2 = 7      # Release order
 
# The Last Jedi
lastJedi = "The Last Jedi"
lastJediNum1 = 8           # Chronological order
lastJediNum2 = 8           # Release order
 
# The Rise of Skywalker
riseOfSkywalker = "The Rise of Skywalker"
riseOfSkywalkerNum1 = 9    # Chronological order
riseOfSkywalkerNum2 = 9    # Release order
 
movies = [
    (phantomMenace,     phantomMenaceNum1,     phantomMenaceNum2),
    (attackOfTheClones, attackOfTheClonesNum1, attackOfTheClonesNum2),
    (revengeOfTheSith,  revengeOfTheSithNum1,  revengeOfTheSithNum2),
    (newHope,           newHopeNum1,           newHopeNum2),
    (empireStrikesBack, empireStrikesBackNum1, empireStrikesBackNum2),
    (returnOfTheJedi,   returnOfTheJediNum1,   returnOfTheJediNum2),
    (forceAwakens,      forceAwakensNum1,      forceAwakensNum2),
    (lastJedi,          lastJediNum1,          lastJediNum2),
    (riseOfSkywalker,   riseOfSkywalkerNum1,   riseOfSkywalkerNum2),
]
 
wb = Workbook()
ws = wb.active
ws.title = "Chronological Order"
ws2 = wb.create_sheet(title="Release Order")
 
gold_fill   = PatternFill("solid", start_color="FFD700", end_color="FFD700")
dark_fill   = PatternFill("solid", start_color="1A1A2E", end_color="1A1A2E")
alt_fill    = PatternFill("solid", start_color="2E2E4E", end_color="2E2E4E")
header_font = Font(name="Arial", bold=True, color="FFD700", size=12)
title_font  = Font(name="Arial", bold=True, color="FFFFFF", size=14)
data_font   = Font(name="Arial", color="FFFFFF", size=11)
center      = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin", color="FFD700"),
    right=Side(style="thin", color="FFD700"),
    top=Side(style="thin", color="FFD700"),
    bottom=Side(style="thin", color="FFD700")
)
 
def style_header(cell, fill=gold_fill, font=None):
    cell.fill = fill
    cell.font = font or Font(name="Arial", bold=True, color="1A1A2E", size=12)
    cell.alignment = center
    cell.border = thin_border
 
def style_data(cell, row_idx):
    cell.fill = dark_fill if row_idx % 2 == 0 else alt_fill
    cell.font = data_font
    cell.alignment = center
    cell.border = thin_border
 
# ── TAB 1: Chronological Order ──
ws.merge_cells("A1:B1")
ws["A1"] = "Chronological Order"
ws["A1"].fill = PatternFill("solid", start_color="4A0000", end_color="4A0000")
ws["A1"].font = title_font
ws["A1"].alignment = center
ws["A1"].border = thin_border
 
ws["A2"] = "#"
ws["B2"] = "Movie Title"
style_header(ws["A2"])
style_header(ws["B2"])
 
chrono_sorted = sorted(movies, key=lambda m: m[1])
for i, (name, chron_num, _) in enumerate(chrono_sorted):
    row = i + 3
    ws.cell(row=row, column=1, value=chron_num)
    ws.cell(row=row, column=2, value=name)
    style_data(ws.cell(row=row, column=1), i)
    style_data(ws.cell(row=row, column=2), i)
 
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 28
for row in range(1, len(movies) + 3):
    ws.row_dimensions[row].height = 24
ws.sheet_view.showGridLines = False
 
# ── TAB 2: Release Order ──
ws2.merge_cells("A1:B1")
ws2["A1"] = "Release Order"
ws2["A1"].fill = PatternFill("solid", start_color="00004A", end_color="00004A")
ws2["A1"].font = title_font
ws2["A1"].alignment = center
ws2["A1"].border = thin_border
 
ws2["A2"] = "#"
ws2["B2"] = "Movie Title"
style_header(ws2["A2"])
style_header(ws2["B2"])
 
release_sorted = sorted(movies, key=lambda m: m[2])
for i, (name, _, rel_num) in enumerate(release_sorted):
    row = i + 3
    ws2.cell(row=row, column=1, value=rel_num)
    ws2.cell(row=row, column=2, value=name)
    style_data(ws2.cell(row=row, column=1), i)
    style_data(ws2.cell(row=row, column=2), i)
 
ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 28
for row in range(1, len(movies) + 3):
    ws2.row_dimensions[row].height = 24
ws2.sheet_view.showGridLines = False

 
wb.save("star_wars_order.xlsx")
 
print("Spreadsheet saved successfully!")
